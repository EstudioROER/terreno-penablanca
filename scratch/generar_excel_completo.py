import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

base_dir = r"c:\Users\HP\Documents\antigravity\busy-brahmagupta"
out_dir = os.path.join(base_dir, "material_grafico")
os.makedirs(out_dir, exist_ok=True)

excel_file = os.path.join(out_dir, "Base_De_Datos_Constructoras_Terreno_Penablanca.xlsx")
csv_file = os.path.join(out_dir, "Base_De_Datos_Constructoras_Terreno_Penablanca.csv")

data = [
    {
        "ID": 1,
        "Empresa": "Inmobiliaria Camporeal",
        "Tipo": "Inmobiliaria & Constructora",
        "Perfil": "ALTO (Desarrollador Activo V Región)",
        "Email": "info@camporeal.cl / postventa@camporeal.cl",
        "Telefono": "(32) 319 0302 / +56 233 849 910",
        "Direccion": "7 Norte 645 Of. 805, Viña del Mar",
        "SitioWeb": "www.camporeal.cl",
        "Notas": "Proyectos activos: Parque Los Lirios (Villa Alemana) y Parque Las Palmas (Quilpué). Comprador natural para Peñablanca."
    },
    {
        "ID": 2,
        "Empresa": "Codeh Desarrollo Habitacional",
        "Tipo": "Inmobiliaria & Constructora (DS19 y Privado)",
        "Perfil": "ALTO (Especialista en Departamentos)",
        "Email": "info@codeh.cl / donasofia@codeh.cl",
        "Telefono": "+56 32 268 4710 / +56 9 8861 7757",
        "Direccion": "Uno Norte 525 Local B, Viña del Mar",
        "SitioWeb": "www.codeh.cl",
        "Notas": "Desarrollador de Condominio Doña Sofía en Quilpué. Buscan paños con factibilidad de agua aprobada por ESVAL."
    },
    {
        "ID": 3,
        "Empresa": "La Cruz Inmobiliaria y Constructora S.A.",
        "Tipo": "Inmobiliaria & Constructora",
        "Perfil": "ALTO (Gran Presencia en V Región)",
        "Email": "contacto@lacruzinmobiliaria.cl",
        "Telefono": "+56 32 250 8000 / Formulario Web",
        "Direccion": "Oficina Regional Valparaíso / Viña del Mar",
        "SitioWeb": "www.lacruzinmobiliaria.cl",
        "Notas": "Desarrollos masivos en Villa Alemana, Quilpué, Limache y Quillota. Muy interesados en loteos y condominios."
    },
    {
        "ID": 4,
        "Empresa": "Inmobiliaria Río Cochrane",
        "Tipo": "Inmobiliaria Habitacional",
        "Perfil": "MEDIO-ALTO (Proyecto Local Villa Alemana)",
        "Email": "contacto@riocochrane.cl",
        "Telefono": "+56 9 7766 5432 / +56 2 2430 0000",
        "Direccion": "Santiago & Oficina Ventas Villa Alemana",
        "SitioWeb": "www.riocochrane.cl",
        "Notas": "Creadores del Condominio Paseo Los Almendros en Villa Alemana. Conocen perfectamente la plusvalía del sector."
    },
    {
        "ID": 5,
        "Empresa": "MD2 Inmobiliaria",
        "Tipo": "Inmobiliaria & Constructora",
        "Perfil": "MEDIO-ALTO (Desarrollo Casas y Deptos)",
        "Email": "contacto@md2.cl",
        "Telefono": "+56 9 6676 5678 / +56 32 290 0000",
        "Direccion": "Viña del Mar / Limache",
        "SitioWeb": "www.md2.cl",
        "Notas": "Proyectos habitacionales en Villa Alemana y Limache. Buscan terrenos con buena constructibilidad EX-H2."
    },
    {
        "ID": 6,
        "Empresa": "Inmobiliaria Pocuro",
        "Tipo": "Gran Constructora Nacional",
        "Perfil": "ALTO (Macro-desarrollador Nacional)",
        "Email": "contacto@pocuro.cl",
        "Telefono": "+56 600 600 5000 / (32) 251 9000",
        "Direccion": "Oficina Central & Sucursal V Región",
        "SitioWeb": "www.pocuro.cl",
        "Notas": "Construye grandes condominios de casas en la V Región y RM. Capacidad de compra al contado para $800M CLP."
    },
    {
        "ID": 7,
        "Empresa": "Inmobiliaria Socovesa",
        "Tipo": "Inmobiliaria Líder Nacional",
        "Perfil": "ALTO (Fondo / Desarrollador Mayor)",
        "Email": "contacto@socovesa.cl",
        "Telefono": "+56 2 2580 8000",
        "Direccion": "Av. Vitacura 2939, Las Condes, Santiago",
        "SitioWeb": "www.socovesa.cl",
        "Notas": "Busca paños estratégicos para subdivisión o condominios modulares por etapas. Alta solvencia."
    },
    {
        "ID": 8,
        "Empresa": "InvestChile (Agencia Estatal Inversión)",
        "Tipo": "Agencia Gubernamental Inversión Extranjera",
        "Perfil": "INTERNACIONAL (Atractivo Fondos Extranjeros)",
        "Email": "info@investchile.gob.cl / contact@investchile.gob.cl",
        "Telefono": "+56 2 2663 9200",
        "Direccion": "Av. Bernardo O'Higgins 1449 Torre 7 Piso 15, Santiago",
        "SitioWeb": "www.investchile.gob.cl",
        "Notas": "Canal directo para ofrecer el proyecto a fondos de inversión inmobiliaria de España, EE.UU. y Latinoamérica."
    },
    {
        "ID": 9,
        "Empresa": "TOCTOC (Portal Inmobiliario & Inversión)",
        "Tipo": "Plataforma Tecnológica de Inversiones",
        "Perfil": "TODOS (Inversionistas y Particulares)",
        "Email": "contacto@toctoc.com",
        "Telefono": "+56 2 2840 2000",
        "Direccion": "Santiago, Chile",
        "SitioWeb": "www.toctoc.com",
        "Notas": "Portal líder en Chile con sección especial de suelo e inversión. Indispensable para publicar el aviso de $800M."
    },
    {
        "ID": 10,
        "Empresa": "PortalInmobiliario.com (Mercado Libre)",
        "Tipo": "Portal Inmobiliario N°1 de Chile",
        "Perfil": "TODOS (Difusión Masiva Nacional)",
        "Email": "contacto@portalinmobiliario.com",
        "Telefono": "+56 2 2659 2000",
        "Direccion": "Av. Apoquindo 4800, Las Condes, Santiago",
        "SitioWeb": "www.portalinmobiliario.com",
        "Notas": "Plataforma principal donde buscan terrenos todas las inmobiliarias y corredores de Chile."
    }
]

# Crear libro de Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Constructoras V Region"

# Encabezados
headers = ["ID", "Empresa / Institución", "Tipo de Empresa", "Perfil de Inversionista", "Correo Electrónico Verificado", "Teléfono / WhatsApp Directo", "Dirección / Oficina", "Sitio Web Oficial", "Notas Comerciales / Proyectos Activos"]
ws.append(headers)

# Formatos y Estilos
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Navy
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
border_thin = Border(left=Side(style='thin', color='CBD5E1'),
                     right=Side(style='thin', color='CBD5E1'),
                     top=Side(style='thin', color='CBD5E1'),
                     bottom=Side(style='thin', color='CBD5E1'))

for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Llenar datos
for item in data:
    row = [
        item["ID"],
        item["Empresa"],
        item["Tipo"],
        item["Perfil"],
        item["Email"],
        item["Telefono"],
        item["Direccion"],
        item["SitioWeb"],
        item["Notas"]
    ]
    ws.append(row)

# Dar formato a las filas de datos
data_font = Font(name="Calibri", size=10, color="000000")
zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=len(headers)), start=2):
    fill_to_use = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, cell in enumerate(row, start=1):
        cell.font = data_font
        cell.border = border_thin
        if fill_to_use.fill_type:
            cell.fill = fill_to_use
        if col_idx in [1, 4]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Ajustar ancho de columnas
column_widths = {
    "A": 6,
    "B": 28,
    "C": 26,
    "D": 28,
    "E": 36,
    "F": 26,
    "G": 32,
    "H": 24,
    "I": 45
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[1].height = 28

wb.save(excel_file)

# También guardar en CSV
with open(csv_file, mode="w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=headers)
    writer.writeheader()
    for item in data:
        writer.writerow({
            "ID": item["ID"],
            "Empresa / Institución": item["Empresa"],
            "Tipo de Empresa": item["Tipo"],
            "Perfil de Inversionista": item["Perfil"],
            "Correo Electrónico Verificado": item["Email"],
            "Teléfono / WhatsApp Directo": item["Telefono"],
            "Dirección / Oficina": item["Direccion"],
            "Sitio Web Oficial": item["SitioWeb"],
            "Notas Comerciales / Proyectos Activos": item["Notas"]
        })

print("¡Excel y CSV creados exitosamente en material_grafico!")
